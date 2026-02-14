# -*- coding: utf-8 -*-
"""
SFC_View: Flask app on port 5556.
User picks start/end datetime -> Apply Filter -> SFC API (with -2h/+2h) -> parse HTML -> filter -> analytics.
Also includes bonepile upload and disposition.
"""
from __future__ import annotations

import io
import os
import sqlite3
import tempfile
import threading
from datetime import datetime
from typing import Any, Dict, Optional, Tuple

from flask import Flask, jsonify, render_template, request, Response, send_file

# Excel export templates (formatting preserved in exported XLSX)
APP_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(APP_DIR, "templates")
SKU_SUMMARY_TEMPLATE_PATH = os.path.join(TEMPLATES_DIR, "SKU_Summary.xlsx")
TRAY_SUMMARY_TEMPLATE_PATH = os.path.join(TEMPLATES_DIR, "Tray_Summary_Template.xlsx")
SKU_DISPO_TEMPLATE_PATH = os.path.join(TEMPLATES_DIR, "SKU_Dispo.xlsx")

from sfc.client import request_fail_result
from sfc.parser import parse_fail_result_html, rows_to_csv
from analytics.compute import compute_all
from analytics.sn_list import compute_sn_list
from analytics.error_stats import compute_error_stats, compute_error_stats_sn_list
from config.app_config import TOP_K_ERRORS_DEFAULT
from bonepile_disposition import (
    ensure_db_ready, RawState, _bonepile_status_payload, _save_uploaded_bonepile_file,
    _copy_for_parse, new_job_id, set_job, run_bonepile_parse_job,
    _load_bonepile_workbook, _find_header_row, _read_header_map, _auto_mapping_from_headers,
    _mapping_errors, _close_and_release_workbook, _remove_temp_file,
    _parse_ca_input_datetime, utc_ms, compute_disposition_stats, compute_disposition_sn_list,
    BONEPILE_UPLOAD_PATH, BONEPILE_IGNORED_SHEETS, ANALYTICS_CACHE_DIR, scan_lock, jobs_lock, jobs
)

app = Flask(__name__)

# Cache last query result for sn-list drill-down
_last_query_lock = threading.Lock()
_last_query_result: Optional[Dict[str, Any]] = None

# Cache last error-stats result for drill-down
_last_error_stats_lock = threading.Lock()
_last_error_stats_result: Optional[Dict[str, Any]] = None


def _parse_datetime(s: Optional[str], is_end: bool = False) -> Optional[datetime]:
    if not s or not s.strip():
        return None
    s = s.strip()
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(s, fmt)
            if fmt == "%Y-%m-%d" and is_end:
                dt = dt.replace(hour=23, minute=59, second=59, microsecond=999999)
            elif fmt == "%Y-%m-%d":
                dt = dt.replace(hour=0, minute=0, second=0, microsecond=0)
            return dt
        except ValueError:
            continue
    return None


def _xlsx_response(data: bytes, filename: str) -> Response:
    resp = Response(data, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.headers["Cache-Control"] = "no-store"
    return resp


def _copy_cell_style(src_cell, tgt_cell):
    """Copy font, fill, alignment, border, number_format from source cell to target cell."""
    if src_cell.has_style:
        tgt_cell.font = src_cell.font.copy()
        tgt_cell.fill = src_cell.fill.copy()
        tgt_cell.alignment = src_cell.alignment.copy()
        tgt_cell.border = src_cell.border.copy()
        tgt_cell.number_format = src_cell.number_format


def _build_export_xlsx(
    export_kind: str,
    computed: Dict[str, Any],
    start_s: str,
    end_s: str,
    start_ca: datetime,
    end_ca: datetime,
) -> Tuple[bytes, str]:
    """
    Build XLSX using templates; preserve template formatting.
    export_kind: 'summary' | 'sku'
    SFC_View uses computed['tray_summary'] for summary (same structure as Bonepile's summary).
    Returns (xlsx_bytes, filename).
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is not installed; cannot export XLSX")

    start_str = start_ca.strftime("%Y-%m-%d %H:%M")
    end_str = end_ca.strftime("%Y-%m-%d %H:%M")
    header_text = f"Testing from {start_str} to {end_str}"

    if export_kind == "summary":
        path = TRAY_SUMMARY_TEMPLATE_PATH
        if not os.path.isfile(path):
            raise FileNotFoundError(f"Tray Summary template not found: {path}")
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        merged_ranges = list(ws.merged_cells.ranges)
        for mr in merged_ranges:
            mr_str = str(mr)
            if 'A1' in mr_str or mr_str.startswith('A1:'):
                ws.unmerge_cells(mr_str)
        from openpyxl.styles import PatternFill
        a1_cell = ws.cell(row=1, column=1)
        a1_fill = a1_cell.fill.copy() if a1_cell.has_style and a1_cell.fill else None
        if a1_cell.has_style:
            a1_cell.fill = PatternFill()
        header_cell = ws.cell(row=1, column=2, value=header_text)
        ws.merge_cells('B1:D1')
        from openpyxl.styles import Font, Alignment
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if a1_fill:
            header_cell.fill = a1_fill
        text_length = len(header_text)
        estimated_width = max(text_length * 1.2 / 3, 15)
        for col in ['B', 'C', 'D']:
            current_width = ws.column_dimensions[col].width if col in ws.column_dimensions and ws.column_dimensions[col].width else 0
            ws.column_dimensions[col].width = max(current_width, estimated_width)

        # SFC_View tray_summary: { tested: {bp, fresh, total}, pass: {...}, fail: {...} }
        t = computed["tray_summary"]
        ws.cell(row=3, column=1, value="TOTAL")
        ws.cell(row=3, column=2, value=t["tested"].get("bp", 0))
        ws.cell(row=3, column=3, value=t["tested"].get("fresh", 0))
        ws.cell(row=3, column=4, value=t["tested"].get("total", 0))
        ws.cell(row=4, column=1, value="PASS")
        ws.cell(row=4, column=2, value=t["pass"].get("bp", 0))
        ws.cell(row=4, column=3, value=t["pass"].get("fresh", 0))
        ws.cell(row=4, column=4, value=t["pass"].get("total", 0))
        ws.cell(row=5, column=1, value="FAIL")
        ws.cell(row=5, column=2, value=t["fail"].get("bp", 0))
        ws.cell(row=5, column=3, value=t["fail"].get("fresh", 0))
        ws.cell(row=5, column=4, value=t["fail"].get("total", 0))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read(), f"summary_{start_s}_to_{end_s}.xlsx"

    if export_kind == "sku":
        path = SKU_SUMMARY_TEMPLATE_PATH
        if not os.path.isfile(path):
            raise FileNotFoundError(f"SKU Summary template not found: {path}")
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        ws.insert_rows(1)
        from openpyxl.styles import PatternFill
        a2_cell = ws.cell(row=2, column=1)
        a2_fill = a2_cell.fill.copy() if a2_cell.has_style and a2_cell.fill else None
        a1_cell = ws.cell(row=1, column=1)
        if a1_cell.has_style:
            a1_cell.fill = PatternFill()
        if a2_cell.has_style:
            a2_cell.fill = PatternFill()
        header_cell = ws.cell(row=1, column=2, value=header_text)
        ws.merge_cells('B1:D1')
        from openpyxl.styles import Font, Alignment
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if a2_fill:
            header_cell.fill = a2_fill
        text_length = len(header_text)
        estimated_width = max(text_length * 1.2 / 3, 15)
        for col in ['B', 'C', 'D']:
            current_width = ws.column_dimensions[col].width if col in ws.column_dimensions and ws.column_dimensions[col].width else 0
            ws.column_dimensions[col].width = max(current_width, estimated_width)

        sku_rows = computed.get("sku_rows") or []
        ws.cell(row=2, column=1, value="SKU")
        ws.cell(row=2, column=2, value="TESTED")
        ws.cell(row=2, column=3, value="PASS")
        ws.cell(row=2, column=4, value="FAIL")
        first_data_row = 3
        template_last_data_row = 5
        template_data_row_count = 4
        for i, r in enumerate(sku_rows):
            row_num = first_data_row + i
            ws.cell(row=row_num, column=1, value=r.get("sku") or "")
            ws.cell(row=row_num, column=2, value=r.get("tested") or 0)
            ws.cell(row=row_num, column=3, value=r.get("pass") or 0)
            ws.cell(row=row_num, column=4, value=r.get("fail") or 0)
            if row_num > template_last_data_row:
                for col in range(1, 5):
                    src_cell = ws.cell(row=first_data_row, column=col)
                    tgt_cell = ws.cell(row=row_num, column=col)
                    _copy_cell_style(src_cell, tgt_cell)
        num_used = len(sku_rows)
        if num_used < template_data_row_count:
            first_unused = first_data_row + num_used
            num_to_delete = template_data_row_count - num_used
            ws.delete_rows(first_unused, num_to_delete)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read(), f"sku_{start_s}_to_{end_s}.xlsx"

    raise ValueError(f"Unsupported export_kind for XLSX: {export_kind}")


def _build_dispo_sku_xlsx(
    dispo_data: Dict[str, Any],
    start_ca: datetime,
    end_ca: datetime,
) -> Tuple[bytes, str]:
    """Build XLSX for Disposition By SKU. Uses SKU_Dispo.xlsx if present, else creates from scratch."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is not installed; cannot export XLSX")

    by_sku = dispo_data.get("by_sku") or []
    start_str = start_ca.strftime("%Y-%m-%d %H:%M")
    end_str = end_ca.strftime("%Y-%m-%d %H:%M")
    date_range_str = f"From {start_str} to {end_str}"

    if os.path.isfile(SKU_DISPO_TEMPLATE_PATH):
        wb = openpyxl.load_workbook(SKU_DISPO_TEMPLATE_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Disposition By SKU"
        ws.cell(row=2, column=1, value="Part Number")
        ws.cell(row=2, column=2, value="Dispositions")
        ws.cell(row=2, column=3, value="Complete")
        ws.cell(row=2, column=4, value="Waiting")

    ws.cell(row=1, column=2, value=date_range_str)
    first_data_row = 3
    for i, row in enumerate(by_sku):
        r = first_data_row + i
        ws.cell(row=r, column=1, value=row.get("sku") or "")
        ws.cell(row=r, column=2, value=row.get("total") or 0)
        ws.cell(row=r, column=3, value=row.get("complete") or 0)
        ws.cell(row=r, column=4, value=row.get("waiting_igs") or 0)
        if r > first_data_row:
            for col in range(1, 5):
                src_cell = ws.cell(row=first_data_row, column=col)
                tgt_cell = ws.cell(row=r, column=col)
                _copy_cell_style(src_cell, tgt_cell)
    for i in range(len(by_sku)):
        r = first_data_row + i
        src_cell = ws.cell(row=r, column=1)
        tgt_cell = ws.cell(row=r, column=5)
        _copy_cell_style(src_cell, tgt_cell)
    for col_idx, col_letter in enumerate(["A", "B", "C", "D", "E"], start=1):
        max_len = 12
        for r in range(1, first_data_row + len(by_sku) + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)) + 1)
        max_len = min(50, max_len)
        current = ws.column_dimensions[col_letter].width if col_letter in ws.column_dimensions and ws.column_dimensions[col_letter].width else 0
        ws.column_dimensions[col_letter].width = max(current, max_len)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    start_s = start_ca.strftime("%Y%m%d_%H%M")
    end_s = end_ca.strftime("%Y%m%d_%H%M")
    return buf.read(), f"Disposition_By_SKU_{start_s}_to_{end_s}.xlsx"


@app.route("/")
def index():
    """Serve analytics dashboard."""
    return render_template("analytics_dashboard.html")


@app.route("/api/query", methods=["POST"])
def api_query():
    """
    Apply Filter: Body { start_datetime, end_datetime, aggregation?: daily|weekly|monthly }
    Calls SFC API, parses HTML, computes analytics. Caches result for sn-list.
    """
    global _last_query_result
    payload = request.json or {}
    start_s = (payload.get("start_datetime") or "").strip()
    end_s = (payload.get("end_datetime") or "").strip()
    aggregation = (payload.get("aggregation") or "daily").strip().lower()
    if aggregation not in ("daily", "weekly", "monthly"):
        aggregation = "daily"

    user_start = _parse_datetime(start_s, is_end=False)
    user_end = _parse_datetime(end_s, is_end=True)
    if user_start is None or user_end is None:
        return jsonify({"error": "start_datetime and end_datetime required (YYYY-MM-DD HH:MM)"}), 400
    if user_end < user_start:
        return jsonify({"error": "end must be after start"}), 400

    ok, html = request_fail_result(user_start, user_end)
    if not ok:
        return jsonify({"error": "SFC API request failed (login or fail_result)"}), 502

    rows = parse_fail_result_html(html, user_start=user_start, user_end=user_end)
    computed = compute_all(rows, aggregation=aggregation)

    with _last_query_lock:
        _last_query_result = computed

    out = {
        "ok": True,
        "summary": computed["summary"],
        "tray_summary": computed["tray_summary"],
        "sku_rows": computed["sku_rows"],
        "breakdown_rows": computed["breakdown_rows"],
        "test_flow": computed["test_flow"],
        "rows": computed["rows"],
    }
    return jsonify(out)


@app.route("/api/sn-list", methods=["POST"])
def api_sn_list():
    """
    Drill-down SN list. Body: { metric, sku?, period?, station?, outcome?, aggregation? }
    Uses last query result from /api/query.
    """
    payload = request.json or {}
    metric = (payload.get("metric") or "total").strip().lower()
    sku = (payload.get("sku") or "").strip() or None
    period = (payload.get("period") or "").strip() or None
    station = (payload.get("station") or "").strip() or None
    outcome = (payload.get("outcome") or "").strip().lower() or None
    if outcome and outcome not in ("pass", "fail"):
        outcome = None
    aggregation = (payload.get("aggregation") or "daily").strip().lower()
    if aggregation not in ("daily", "weekly", "monthly"):
        aggregation = "daily"

    with _last_query_lock:
        computed = _last_query_result

    if not computed:
        return jsonify({"error": "Apply filter first", "count": 0, "rows": []}), 400

    try:
        rows = compute_sn_list(
            computed,
            metric=metric,
            sku=sku,
            period=period,
            station=station,
            outcome=outcome,
            aggregation=aggregation,
        )
        return jsonify({"ok": True, "count": len(rows), "rows": rows})
    except Exception as e:
        return jsonify({"error": str(e), "count": 0, "rows": []}), 500


@app.route("/api/error-stats", methods=["POST"])
def api_error_stats():
    """
    Failure-focused error statistics for Tray testing.
    Body: { start_datetime, end_datetime, top_k_errors?: 5 }
    Returns tables A-G.
    """
    global _last_error_stats_result
    payload = request.json or {}
    start_s = (payload.get("start_datetime") or "").strip()
    end_s = (payload.get("end_datetime") or "").strip()
    top_k = payload.get("top_k_errors")
    if top_k is None:
        top_k = TOP_K_ERRORS_DEFAULT
    try:
        top_k = int(top_k)
        if top_k < 1:
            top_k = TOP_K_ERRORS_DEFAULT
    except (TypeError, ValueError):
        top_k = TOP_K_ERRORS_DEFAULT

    user_start = _parse_datetime(start_s, is_end=False)
    user_end = _parse_datetime(end_s, is_end=True)
    if user_start is None or user_end is None:
        return jsonify({"error": "start_datetime and end_datetime required (YYYY-MM-DD HH:MM)"}), 400
    if user_end < user_start:
        return jsonify({"error": "end must be after start"}), 400

    ok, html = request_fail_result(user_start, user_end)
    if not ok:
        return jsonify({"error": "SFC API request failed (login or fail_result)"}), 502

    rows = parse_fail_result_html(html, user_start=user_start, user_end=user_end)
    result = compute_error_stats(rows, top_k=top_k)

    with _last_error_stats_lock:
        _last_error_stats_result = result

    fail_rows = result.get("_fail_rows") or []
    total_fail_events = len(fail_rows)
    total_unique_trays = len(
        set(
            (r.get("serial_number") or r.get("sn") or "").strip()
            for r in fail_rows
            if (r.get("serial_number") or r.get("sn") or "").strip()
        )
    )
    top_k_errors = result["top_k_errors"]
    top_3_errors = [e["error_code"] for e in top_k_errors[:3]]
    fail_by_station = result["fail_by_station"]
    top_station = (
        max(fail_by_station, key=lambda x: x.get("fail_events", 0))
        if fail_by_station
        else {"station_group": "-", "fail_events": 0}
    )
    out = {
        "ok": True,
        "total_fail_events": total_fail_events,
        "total_unique_trays": total_unique_trays,
        "top_3_errors": top_3_errors,
        "top_station": {"station_group": top_station["station_group"], "fail_events": top_station["fail_events"]},
        "fail_by_station": result["fail_by_station"],
        "top_k_errors": result["top_k_errors"],
        "station_error_matrix": result["station_error_matrix"],
        "station_error_matrix_cols": result["station_error_matrix_cols"],
        "station_instance_hotspots": result["station_instance_hotspots"],
        "ttc_overall": result["ttc_overall"],
        "ttc_by_station": result["ttc_by_station"],
        "ttc_by_error": result["ttc_by_error"],
    }
    return jsonify(out)


@app.route("/api/error-stats-sn-list", methods=["POST"])
def api_error_stats_sn_list():
    """
    Drill-down SN list for Error Stats.
    Body: { start_datetime, end_datetime, metric, station_group?, error_code?, ttc_bucket?, station_instance? }
    """
    payload = request.json or {}
    metric = (payload.get("metric") or "").strip()
    station_group = (payload.get("station_group") or "").strip() or None
    error_code = (payload.get("error_code") or "").strip() or None
    ttc_bucket = (payload.get("ttc_bucket") or "").strip() or None
    station_instance = (payload.get("station_instance") or "").strip() or None
    drill_type = (payload.get("drill_type") or "").strip() or None

    with _last_error_stats_lock:
        result = _last_error_stats_result

    if not result:
        return jsonify({"error": "Apply filter and load Error Stats first", "count": 0, "rows": []}), 400

    try:
        rows = compute_error_stats_sn_list(
            result,
            metric=metric,
            station_group=station_group,
            error_code=error_code,
            ttc_bucket=ttc_bucket,
            station_instance=station_instance,
            drill_type=drill_type,
        )
        return jsonify({"ok": True, "count": len(rows), "rows": rows})
    except Exception as e:
        return jsonify({"error": str(e), "count": 0, "rows": []}), 500


@app.route("/api/fail_result", methods=["POST"])
def api_fail_result():
    """Legacy: returns rows + csv. Prefer /api/query."""
    payload = request.json or {}
    start_s = (payload.get("start_datetime") or "").strip()
    end_s = (payload.get("end_datetime") or "").strip()

    user_start = _parse_datetime(start_s, is_end=False)
    user_end = _parse_datetime(end_s, is_end=True)
    if user_start is None or user_end is None:
        return jsonify({"error": "start_datetime and end_datetime required (YYYY-MM-DD HH:MM)"}), 400
    if user_end < user_start:
        return jsonify({"error": "end must be after start"}), 400

    ok, html = request_fail_result(user_start, user_end)
    if not ok:
        return jsonify({"error": "SFC API request failed (login or fail_result)"}), 502

    rows = parse_fail_result_html(html, user_start=user_start, user_end=user_end)
    csv_str = rows_to_csv(rows, include_bp=False)

    return jsonify({
        "ok": True,
        "rows": rows,
        "csv": csv_str,
        "count": len(rows),
    })


def _error_stats_to_csv(result: Dict[str, Any]) -> str:
    """Build CSV string from error stats result (multiple tables)."""
    import csv as csv_mod
    buf = io.StringIO(newline="")
    w = csv_mod.writer(buf)

    def write_section(title: str, rows: list, fieldnames: list):
        w.writerow([title])
        w.writerow(fieldnames)
        for r in rows:
            w.writerow([r.get(f, "") for f in fieldnames])
        w.writerow([])

    write_section("Fail by Station", result.get("fail_by_station", []),
                  ["station_group", "fail_events", "unique_tray", "pct_fail_events"])
    write_section("Top K Errors", result.get("top_k_errors", []),
                  ["error_code", "representative_error_message", "fail_events", "unique_tray", "top_station_group"])

    write_section("Station Instance Hotspots", result.get("station_instance_hotspots", []),
                  ["station_instance", "station_group", "fail_events", "unique_tray", "top_error_code"])

    ttc = result.get("ttc_overall", {})
    if ttc:
        skip_keys = {"bucket_leq5m", "bucket_5_15m", "bucket_15_60m", "bucket_gt60m"}
        ttc_filtered = {k: v for k, v in ttc.items() if k not in skip_keys}
        if ttc_filtered:
            w.writerow(["TTC Overall"])
            w.writerow(list(ttc_filtered.keys()))
            w.writerow(list(ttc_filtered.values()))
            w.writerow([])

    write_section("TTC by Station", result.get("ttc_by_station", []),
                  ["station_group", "resolved_count", "open_count", "median_ttc", "mean_ttc", "max_ttc", "total_ttc_minutes"])
    write_section("TTC by Error", result.get("ttc_by_error", []),
                  ["error_code", "resolved_count", "median_ttc", "total_ttc_minutes"])

    return buf.getvalue()


@app.route("/api/export", methods=["POST"])
@app.route("/api/export/csv", methods=["POST"])
def api_export_csv():
    """
    Export CSV or XLSX.
    Body: { start_datetime, end_datetime, aggregation?, export?, format? }
    format=xlsx with export=summary|sku|disposition_by_sku returns Excel.
    Otherwise returns CSV.
    """
    payload = request.json or {}
    start_s = (payload.get("start_datetime") or "").strip()
    end_s = (payload.get("end_datetime") or "").strip()
    export_kind = (payload.get("export") or "dashboard").strip().lower()
    export_format = (payload.get("format") or "csv").strip().lower()

    # Handle XLSX export for summary, sku, disposition_by_sku
    if export_format == "xlsx" and export_kind in ("summary", "sku", "disposition_by_sku"):
        if export_kind == "disposition_by_sku":
            start_ca = _parse_ca_input_datetime(start_s, is_end=False) if start_s else None
            end_ca = _parse_ca_input_datetime(end_s, is_end=True) if end_s else None
            if start_ca is None or end_ca is None:
                return jsonify({"error": "start_datetime and end_datetime required for disposition export"}), 400
            if end_ca <= start_ca:
                return jsonify({"error": "end must be after start"}), 400
            try:
                ensure_db_ready()
                start_ca_ms = utc_ms(start_ca)
                end_ca_ms = utc_ms(end_ca)
                dispo_data = compute_disposition_stats(aggregation="daily", start_ca_ms=start_ca_ms, end_ca_ms=end_ca_ms)
                data_xlsx, filename = _build_dispo_sku_xlsx(dispo_data, start_ca, end_ca)
                return _xlsx_response(data_xlsx, filename)
            except sqlite3.OperationalError:
                dispo_data = {"by_sku": []}
                data_xlsx, filename = _build_dispo_sku_xlsx(dispo_data, start_ca, end_ca)
                return _xlsx_response(data_xlsx, filename)
            except FileNotFoundError as e:
                return jsonify({"error": str(e)}), 404
            except Exception as e:
                return jsonify({"error": f"XLSX export failed: {str(e)}"}), 500
        else:
            user_start = _parse_datetime(start_s, is_end=False)
            user_end = _parse_datetime(end_s, is_end=True)
            if user_start is None or user_end is None:
                return jsonify({"error": "start_datetime and end_datetime required"}), 400
            if user_end < user_start:
                return jsonify({"error": "end must be after start"}), 400
            aggregation = (payload.get("aggregation") or "daily").strip().lower()
            if aggregation not in ("daily", "weekly", "monthly"):
                aggregation = "daily"
            ok, html = request_fail_result(user_start, user_end)
            if not ok:
                return jsonify({"error": "SFC API request failed"}), 502
            rows = parse_fail_result_html(html, user_start=user_start, user_end=user_end)
            computed = compute_all(rows, aggregation=aggregation)
            start_fmt = user_start.strftime("%Y%m%d_%H%M")
            end_fmt = user_end.strftime("%Y%m%d_%H%M")
            try:
                data_xlsx, filename = _build_export_xlsx(
                    export_kind, computed, start_fmt, end_fmt, user_start, user_end
                )
                return _xlsx_response(data_xlsx, filename)
            except FileNotFoundError as e:
                return jsonify({"error": str(e)}), 404
            except Exception as e:
                return jsonify({"error": f"XLSX export failed: {str(e)}"}), 500

    # CSV export - error_stats
    if export_kind == "error_stats":
        top_k = payload.get("top_k_errors")
        if top_k is None:
            top_k = TOP_K_ERRORS_DEFAULT
        try:
            top_k = int(top_k)
            if top_k < 1:
                top_k = TOP_K_ERRORS_DEFAULT
        except (TypeError, ValueError):
            top_k = TOP_K_ERRORS_DEFAULT

        user_start = _parse_datetime(start_s, is_end=False)
        user_end = _parse_datetime(end_s, is_end=True)
        if user_start is None or user_end is None:
            return jsonify({"error": "start_datetime and end_datetime required"}), 400
        if user_end < user_start:
            return jsonify({"error": "end must be after start"}), 400

        ok, html = request_fail_result(user_start, user_end)
        if not ok:
            return jsonify({"error": "SFC API request failed"}), 502

        rows = parse_fail_result_html(html, user_start=user_start, user_end=user_end)
        result = compute_error_stats(rows, top_k=top_k)
        csv_str = _error_stats_to_csv(result)
        filename = f"error_stats_{user_start.strftime('%Y%m%d_%H%M')}_to_{user_end.strftime('%Y%m%d_%H%M')}.csv"
        buf = io.BytesIO(csv_str.encode("utf-8-sig"))
        return send_file(
            buf,
            mimetype="text/csv; charset=utf-8-sig",
            as_attachment=True,
            download_name=filename,
        )

    # CSV export - default fail_result
    user_start = _parse_datetime(start_s, is_end=False)
    user_end = _parse_datetime(end_s, is_end=True)
    if user_start is None or user_end is None:
        return jsonify({"error": "start_datetime and end_datetime required"}), 400
    if user_end < user_start:
        return jsonify({"error": "end must be after start"}), 400

    ok, html = request_fail_result(user_start, user_end)
    if not ok:
        return jsonify({"error": "SFC API request failed"}), 502

    rows = parse_fail_result_html(html, user_start=user_start, user_end=user_end)
    from analytics.bp_check import add_bp_to_rows
    rows = add_bp_to_rows(rows)
    csv_str = rows_to_csv(rows, include_bp=True)
    filename = f"fail_result_{user_start.strftime('%Y%m%d_%H%M')}_to_{user_end.strftime('%Y%m%d_%H%M')}.csv"

    buf = io.BytesIO(csv_str.encode("utf-8-sig"))
    return send_file(
        buf,
        mimetype="text/csv; charset=utf-8-sig",
        as_attachment=True,
        download_name=filename,
    )


# -----------------------------
# Bonepile Upload & Disposition Routes (copied from Bonepile_view)
# -----------------------------


@app.route("/api/job/<job_id>")
def api_job(job_id: str):
    """Get job status."""
    with jobs_lock:
        data = jobs.get(job_id)
    if not data:
        return jsonify({"error": "job not found"}), 404
    return jsonify(data)


@app.route("/api/bonepile/status")
def api_bonepile_status():
    """Get bonepile upload/parse status."""
    ensure_db_ready()
    state = RawState.load()
    return jsonify(_bonepile_status_payload(state))


@app.route("/api/bonepile/upload", methods=["POST"])
def api_bonepile_upload():
    """
    Upload NV/IGS workbook. The backend stores only the latest file (replaces previous).
    After upload, automatically parse all allowed sheets with auto-detect.
    """
    try:
        ensure_db_ready()
        try:
            import openpyxl
        except ImportError:
            openpyxl = None
        if openpyxl is None:
            return jsonify({"error": "openpyxl not installed; cannot accept XLSX"}), 500
        if "file" not in request.files:
            return jsonify({"error": "file is required"}), 400
        f = request.files["file"]
        if not f or not getattr(f, "filename", ""):
            return jsonify({"error": "no file selected"}), 400
        name = str(f.filename)
        if not name.lower().endswith(".xlsx"):
            return jsonify({"error": "only .xlsx is supported for bonepile upload"}), 400

        with scan_lock:
            state = RawState.load()
            meta = _save_uploaded_bonepile_file(f)
            state.bonepile_file = meta
            state.bonepile_sheet_status = state.bonepile_sheet_status or {}
            state.save()

        job_id = new_job_id()
        parse_copy = os.path.join(ANALYTICS_CACHE_DIR, "bonepile_parse_" + job_id + ".xlsx")
        if not _copy_for_parse(parse_copy):
            return jsonify({"error": "Upload ok but could not start parse (file in use?)"}), 500
        set_job(job_id, status="queued", message="Auto-parsing all sheets with auto-detect...")
        t = threading.Thread(target=run_bonepile_parse_job, args=(job_id, None), kwargs={"path": parse_copy}, daemon=True)
        t.start()
        return jsonify({"ok": True, "job_id": job_id, "bonepile_file": meta})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/bonepile/sheets")
def api_bonepile_sheets():
    """Return sheet list."""
    try:
        import openpyxl
    except ImportError:
        openpyxl = None
    if openpyxl is None:
        return jsonify({"error": "openpyxl not installed; cannot read XLSX"}), 500
    state = RawState.load()
    if not os.path.exists(BONEPILE_UPLOAD_PATH):
        return jsonify({"ok": True, "has_file": False, "ignored": BONEPILE_IGNORED_SHEETS, "sheets": {}})
    import shutil
    fd, copy_path = tempfile.mkstemp(suffix=".xlsx", prefix="bonepile_sheets_", dir=ANALYTICS_CACHE_DIR)
    os.close(fd)
    wb = None
    try:
        shutil.copy2(BONEPILE_UPLOAD_PATH, copy_path)
        wb = _load_bonepile_workbook(copy_path)
        all_sheets = list(wb.sheetnames)
        ignored = [s for s in all_sheets if s in BONEPILE_IGNORED_SHEETS]
        processable = [s for s in all_sheets if s not in BONEPILE_IGNORED_SHEETS]
        out: dict = {}
        for sheet in processable:
            if sheet not in all_sheets:
                out[sheet] = {"present": False}
                continue
            ws = wb[sheet]
            header_row = _find_header_row(ws) or 0
            header_map = _read_header_map(ws, header_row) if header_row else {}
            auto_map = _auto_mapping_from_headers(header_map) if header_map else {}
            errs = _mapping_errors(auto_map, header_map) if header_map else ["Header row not found (SN)"]
            out[sheet] = {
                "present": True,
                "header_row": int(header_row) if header_row else None,
                "headers": list(header_map.keys())[:80],
                "auto_columns": auto_map,
                "auto_errors": errs,
                "saved_mapping": (state.bonepile_mapping or {}).get(sheet),
                "status": (state.bonepile_sheet_status or {}).get(sheet),
            }
        return jsonify({"ok": True, "has_file": True, "ignored": ignored, "sheets": out})
    except Exception as e:
        return jsonify({"error": "Failed to read workbook: " + str(e)}), 500
    finally:
        if wb is not None:
            _close_and_release_workbook(wb)
            wb = None
        _remove_temp_file(copy_path)


@app.route("/api/bonepile/mapping", methods=["POST"])
def api_bonepile_mapping():
    """
    Save mapping for a single sheet:
      { sheet, header_row, columns: {sn, nv_disposition, status, pic, igs_action, igs_status, nvpn?} }
    """
    payload = request.json or {}
    sheet = str(payload.get("sheet") or "").strip()
    if sheet in BONEPILE_IGNORED_SHEETS:
        return jsonify({"error": "sheet is in ignored list"}), 400
    header_row = int(payload.get("header_row") or 0)
    columns = payload.get("columns") if isinstance(payload.get("columns"), dict) else {}
    if header_row <= 0:
        return jsonify({"error": "header_row must be >= 1"}), 400

    with scan_lock:
        state = RawState.load()
        if state.bonepile_mapping is None:
            state.bonepile_mapping = {}
        state.bonepile_mapping[sheet] = {"header_row": int(header_row), "columns": columns}
        state.save()

    job_id = new_job_id()
    parse_copy = os.path.join(ANALYTICS_CACHE_DIR, "bonepile_parse_" + job_id + ".xlsx")
    if not os.path.exists(BONEPILE_UPLOAD_PATH) or not _copy_for_parse(parse_copy):
        return jsonify({"error": "Could not copy workbook for parse (file missing or in use)"}), 500
    set_job(job_id, status="queued", message=f"Parsing {sheet}...")
    t = threading.Thread(target=run_bonepile_parse_job, args=(job_id, [sheet]), kwargs={"path": parse_copy}, daemon=True)
    t.start()
    return jsonify({"ok": True, "job_id": job_id})


@app.route("/api/bonepile/parse", methods=["POST"])
def api_bonepile_parse():
    """Trigger parse job (all sheets or a single sheet)."""
    ensure_db_ready()
    payload = request.json or {}
    sheet = str(payload.get("sheet") or "").strip() if payload.get("sheet") is not None else ""
    sheets: Optional[list] = None
    if sheet:
        if sheet in BONEPILE_IGNORED_SHEETS:
            return jsonify({"error": "sheet is in ignored list"}), 400
        sheets = [sheet]
    job_id = new_job_id()
    parse_copy = os.path.join(ANALYTICS_CACHE_DIR, "bonepile_parse_" + job_id + ".xlsx")
    if not os.path.exists(BONEPILE_UPLOAD_PATH) or not _copy_for_parse(parse_copy):
        return jsonify({"error": "Could not copy workbook for parse (file missing or in use)"}), 500
    set_job(job_id, status="queued", message="Bonepile parse queued")
    t = threading.Thread(target=run_bonepile_parse_job, args=(job_id, sheets), kwargs={"path": parse_copy}, daemon=True)
    t.start()
    return jsonify({"ok": True, "job_id": job_id})


@app.route("/api/bonepile/disposition")
def api_bonepile_disposition():
    """
    NV Disposition stats from bonepile_entries.
    Query: aggregation=daily|weekly|monthly (default daily), start_datetime, end_datetime (optional).
    Returns: summary { total, waiting_igs, complete }, by_sku, by_period.
    """
    ensure_db_ready()
    aggregation = request.args.get("aggregation", "daily").strip().lower()
    if aggregation not in ("daily", "weekly", "monthly"):
        aggregation = "daily"
    start_dt = request.args.get("start_datetime")
    end_dt = request.args.get("end_datetime")
    start_ca_ms = None
    end_ca_ms = None
    if start_dt:
        start_ca = _parse_ca_input_datetime(start_dt, is_end=False)
        if start_ca:
            start_ca_ms = utc_ms(start_ca)
    if end_dt:
        end_ca = _parse_ca_input_datetime(end_dt, is_end=True)
        if end_ca:
            end_ca_ms = utc_ms(end_ca)
    try:
        data = compute_disposition_stats(aggregation=aggregation, start_ca_ms=start_ca_ms, end_ca_ms=end_ca_ms)
        return jsonify({"ok": True, **data})
    except sqlite3.OperationalError:
        return jsonify({
            "ok": True,
            "summary": {"total": 0, "waiting_igs": 0, "complete": 0, "unique_trays_bp": 0, "all_pass_trays": 0},
            "by_sku": [],
            "by_period": [],
            "tray_by_sku": [],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/bonepile/disposition/sn-list", methods=["POST"])
def api_bonepile_disposition_sn_list():
    """
    SN list for disposition drill-down.
    Body: { metric: total|waiting|complete, sku?: string, period?: string, aggregation?: daily|weekly|monthly, start_datetime?, end_datetime? }.
    """
    ensure_db_ready()
    payload = request.json or {}
    metric = str(payload.get("metric") or "total").strip().lower()
    if metric not in ("total", "waiting", "complete", "trays_bp", "all_pass_trays"):
        metric = "total"
    sku = (payload.get("sku") or "").strip() or None
    period = (payload.get("period") or "").strip() or None
    aggregation = str(payload.get("aggregation") or "daily").strip().lower()
    if aggregation not in ("daily", "weekly", "monthly"):
        aggregation = "daily"
    start_dt = payload.get("start_datetime")
    end_dt = payload.get("end_datetime")
    start_ca_ms = None
    end_ca_ms = None
    if start_dt:
        start_ca = _parse_ca_input_datetime(start_dt, is_end=False)
        if start_ca:
            start_ca_ms = utc_ms(start_ca)
    if end_dt:
        end_ca = _parse_ca_input_datetime(end_dt, is_end=True)
        if end_ca:
            end_ca_ms = utc_ms(end_ca)
    try:
        rows = compute_disposition_sn_list(metric=metric, sku=sku, period=period, aggregation=aggregation, start_ca_ms=start_ca_ms, end_ca_ms=end_ca_ms)
        return jsonify({"ok": True, "count": len(rows), "rows": rows})
    except sqlite3.OperationalError:
        return jsonify({"ok": True, "count": 0, "rows": []})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5556, debug=True)
