# -*- coding: utf-8 -*-
"""
Bonepile upload and disposition logic (copied from Bonepile_view/analytics_server.py).
Handles NV/IGS workbook upload, parsing, and disposition stats computation.
"""
from __future__ import annotations

import hashlib
import json
import os
import re
import sqlite3
import threading
import time
import gc
import shutil
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

import pytz

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None

# Config
APP_DIR = os.path.dirname(os.path.abspath(__file__))
ANALYTICS_CACHE_DIR = os.path.join(APP_DIR, "analytics_cache")
DB_PATH = os.path.join(ANALYTICS_CACHE_DIR, "analytics.db")
STATE_PATH = os.path.join(ANALYTICS_CACHE_DIR, "raw_state.json")

# Uploaded NV/IGS bonepile workbook (single file; replaced on each upload)
BONEPILE_UPLOAD_PATH = os.path.join(ANALYTICS_CACHE_DIR, "bonepile_upload.xlsx")
# Cache of BP SNs from NV disposition sheets: only add, never delete (used for is_bonepile)
BP_SN_CACHE_PATH = os.path.join(ANALYTICS_CACHE_DIR, "bp_sn_cache.json")
# Sheets to process (block-list style: only these are allowed; all others ignored)
BONEPILE_ALLOWED_SHEETS = ["TS2-SKU1100", "VR-TS1", "TS2-SKU002", "TS2-SKU010"]
BONEPILE_REQUIRED_FIELDS = ["sn", "nv_disposition", "status", "pic", "igs_action", "igs_status"]

CA_TZ = pytz.timezone("America/Los_Angeles")

# BP SN cache: load set (read-only) and merge new SNs (never delete)
_bp_sn_cache_lock = threading.Lock()
_bp_sn_cache_set: Optional[Set[str]] = None

# Job system (in-memory)
jobs_lock = threading.Lock()
jobs: Dict[str, Dict[str, Any]] = {}

scan_lock = threading.Lock()

db_init_lock = threading.Lock()
db_initialized = False


def ensure_dirs() -> None:
    os.makedirs(ANALYTICS_CACHE_DIR, exist_ok=True)


def utc_ms(dt: datetime) -> int:
    """Convert aware datetime to epoch milliseconds."""
    if dt.tzinfo is None:
        raise ValueError("utc_ms expects tz-aware datetime")
    return int(dt.timestamp() * 1000)


def load_bp_sn_set() -> Set[str]:
    """Load set of BP SNs from cache file. Returns empty set if file missing or invalid."""
    global _bp_sn_cache_set
    with _bp_sn_cache_lock:
        if _bp_sn_cache_set is not None:
            return set(_bp_sn_cache_set)
    try:
        if not os.path.isfile(BP_SN_CACHE_PATH):
            return set()
        with open(BP_SN_CACHE_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        sns = data.get("sns") if isinstance(data, dict) else None
        if not isinstance(sns, list):
            return set()
        out = set(str(s).strip() for s in sns if s and str(s).strip())
    except Exception:
        out = set()
    with _bp_sn_cache_lock:
        _bp_sn_cache_set = out
    return set(out)


def update_bp_sn_cache(new_sns: Iterable[str]) -> None:
    """Merge new_sns into BP cache file. Never removes existing SNs."""
    global _bp_sn_cache_set
    new_set = set(str(s).strip() for s in new_sns if s and str(s).strip())
    if not new_set:
        return
    os.makedirs(ANALYTICS_CACHE_DIR, exist_ok=True)
    with _bp_sn_cache_lock:
        current: Set[str] = set()
        if os.path.isfile(BP_SN_CACHE_PATH):
            try:
                with open(BP_SN_CACHE_PATH, "r", encoding="utf-8") as f:
                    data = json.load(f)
                sns = data.get("sns") if isinstance(data, dict) else None
                if isinstance(sns, list):
                    current = set(str(s).strip() for s in sns if s and str(s).strip())
            except Exception:
                pass
        current.update(new_set)
        _bp_sn_cache_set = set(current)
        data = {"sns": sorted(current), "last_updated_ca_ms": int(time.time() * 1000)}
        try:
            with open(BP_SN_CACHE_PATH, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=0)
        except Exception:
            pass


def invalidate_bp_sn_cache() -> None:
    """Clear in-memory cache so next load_bp_sn_set() reads from file."""
    global _bp_sn_cache_set
    with _bp_sn_cache_lock:
        _bp_sn_cache_set = None


@dataclass
class RawState:
    min_ca_ms: Optional[int] = None
    max_ca_ms: Optional[int] = None
    min_key: Optional[Tuple[int, str]] = None
    max_key: Optional[Tuple[int, str]] = None
    min_path: Optional[str] = None
    max_path: Optional[str] = None
    last_scan_ca_ms: Optional[int] = None
    full_day_runs: Optional[Dict[str, str]] = None
    scanned_tw_dates: Optional[List[str]] = None
    bonepile_file: Optional[Dict[str, Any]] = None
    bonepile_mapping: Optional[Dict[str, Any]] = None
    bonepile_sheet_status: Optional[Dict[str, Any]] = None

    @staticmethod
    def load() -> "RawState":
        if not os.path.exists(STATE_PATH):
            return RawState()
        try:
            with open(STATE_PATH, "r", encoding="utf-8") as f:
                data = json.load(f) or {}
            return RawState(
                min_ca_ms=data.get("min_ca_ms"),
                max_ca_ms=data.get("max_ca_ms"),
                min_key=tuple(data["min_key"]) if isinstance(data.get("min_key"), list) else None,
                max_key=tuple(data["max_key"]) if isinstance(data.get("max_key"), list) else None,
                min_path=data.get("min_path"),
                max_path=data.get("max_path"),
                last_scan_ca_ms=data.get("last_scan_ca_ms"),
                full_day_runs=data.get("full_day_runs") if isinstance(data.get("full_day_runs"), dict) else None,
                scanned_tw_dates=data.get("scanned_tw_dates") if isinstance(data.get("scanned_tw_dates"), list) else None,
                bonepile_file=data.get("bonepile_file") if isinstance(data.get("bonepile_file"), dict) else None,
                bonepile_mapping=data.get("bonepile_mapping") if isinstance(data.get("bonepile_mapping"), dict) else None,
                bonepile_sheet_status=data.get("bonepile_sheet_status")
                if isinstance(data.get("bonepile_sheet_status"), dict)
                else None,
            )
        except Exception:
            return RawState()

    def save(self) -> None:
        tmp = STATE_PATH + ".tmp"
        data = {
            "min_ca_ms": self.min_ca_ms,
            "max_ca_ms": self.max_ca_ms,
            "min_key": list(self.min_key) if self.min_key else None,
            "max_key": list(self.max_key) if self.max_key else None,
            "min_path": self.min_path,
            "max_path": self.max_path,
            "last_scan_ca_ms": self.last_scan_ca_ms,
            "full_day_runs": self.full_day_runs or None,
            "scanned_tw_dates": self.scanned_tw_dates or None,
            "bonepile_file": self.bonepile_file or None,
            "bonepile_mapping": self.bonepile_mapping or None,
            "bonepile_sheet_status": self.bonepile_sheet_status or None,
        }
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        if os.path.exists(STATE_PATH):
            os.remove(STATE_PATH)
        os.replace(tmp, STATE_PATH)


def ensure_db_ready(force: bool = False) -> None:
    """Ensure analytics cache directory + SQLite schema exist."""
    global db_initialized
    with db_init_lock:
        if db_initialized and not force:
            return
        ensure_dirs()
        init_db()
        db_initialized = True


def connect_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );
            """
        )
        conn.commit()

        # NV/IGS workbook parsed rows (per sheet)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS bonepile_entries (
              sheet TEXT NOT NULL,
              excel_row INTEGER NOT NULL,
              sn TEXT NOT NULL,
              nvpn TEXT,
              status TEXT,
              pic TEXT,
              igs_status TEXT,
              nv_disposition TEXT,
              igs_action TEXT,
              nv_dispo_count INTEGER,
              igs_action_count INTEGER,
              updated_at_ca_ms INTEGER,
              PRIMARY KEY (sheet, excel_row)
            );
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bp_sn ON bonepile_entries (sn);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bp_sheet_sn ON bonepile_entries (sheet, sn);")
        conn.commit()
    finally:
        conn.close()


def new_job_id() -> str:
    return f"job_{int(time.time() * 1000)}_{os.getpid()}"


def set_job(job_id: str, **fields: Any) -> None:
    with jobs_lock:
        jobs.setdefault(job_id, {})
        jobs[job_id].update(fields)


def _normalize_sn(val: Any) -> Optional[str]:
    """Normalize SN value to 13-digit string starting with 18, or None."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # Handle Excel scientific notation
    try:
        if re.fullmatch(r"\d+(\.\d+)?E\+\d+", s, flags=re.IGNORECASE):
            s = str(int(float(s)))
    except Exception:
        pass
    # Strip trailing .0
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    s = re.sub(r"[^\d]", "", s)
    if len(s) == 13 and s.startswith("18"):
        return s
    return None


def _extract_mmdd_entries(text: Any) -> List[str]:
    """Extract "entries" from a cell that may contain multiple mm/dd markers."""
    if text is None:
        return []
    raw = str(text)
    if not raw.strip():
        return []
    matches = list(re.finditer(r"\b\d{1,2}/\d{1,2}\b", raw))
    if not matches:
        return []
    out: List[str] = []
    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(raw)
        seg = raw[start:end].strip()
        if seg:
            out.append(seg)
    return out


def _load_bonepile_workbook(path: str):
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed; cannot read XLSX")
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _hash_sheet_content(ws, max_rows: int = 10000) -> str:
    """Compute SHA256 hash of sheet content."""
    h = hashlib.sha256()
    row_count = 0
    for row in ws.iter_rows(max_row=max_rows, values_only=True):
        if row_count >= max_rows:
            break
        row_str = "|".join(str(v if v is not None else "") for v in row)
        h.update(row_str.encode("utf-8"))
        h.update(b"\n")
        row_count += 1
    h.update(str(row_count).encode("utf-8"))
    return h.hexdigest()


def _find_header_row(ws, max_rows: int = 300) -> Optional[int]:
    """Return 1-based row index of header row containing 'SN' cell."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        if not row:
            continue
        for v in row:
            if v is None:
                continue
            if str(v).strip().upper() == "SN":
                return i
    return None


def _read_header_map(ws, header_row: int, max_cols: int = 80) -> Dict[str, int]:
    """Build case-insensitive header -> 1-based column index map."""
    header_map: Dict[str, int] = {}
    for j, cell in enumerate(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)), start=1):
        if j > max_cols:
            break
        if cell is None:
            continue
        name = str(cell).strip()
        if not name:
            continue
        header_map[name.strip().upper()] = j
    return header_map


def _auto_mapping_from_headers(header_map: Dict[str, int]) -> Dict[str, int]:
    """Auto-map required fields by header names."""
    def pick(*names: str) -> Optional[int]:
        for n in names:
            idx = header_map.get(n.upper())
            if idx:
                return idx
        return None

    m: Dict[str, int] = {}
    m["sn"] = pick("SN") or 0
    m["nv_disposition"] = pick("NV DISPOSITION", "NV DISPO", "NV DISPOSITION ") or 0
    m["status"] = pick("STATUS") or 0
    m["pic"] = pick("PIC") or 0
    m["igs_action"] = pick("IGS ACTION") or 0
    m["igs_status"] = pick("IGS STATUS") or 0
    m["nvpn"] = pick("NVPN", "PART NUMBER", "PART NUMBERS", "SKU") or 0
    return m


def _mapping_errors(mapping: Dict[str, int], header_map: Dict[str, int]) -> List[str]:
    errors: List[str] = []
    for k in BONEPILE_REQUIRED_FIELDS:
        if int(mapping.get(k) or 0) <= 0:
            errors.append(f"Missing column for '{k}'")
    if errors:
        sample = ", ".join(list(header_map.keys())[:25])
        errors.append(f"Available headers: {sample}")
    return errors


def _close_and_release_workbook(wb) -> None:
    """Close workbook and release file handle."""
    if wb is None:
        return
    try:
        wb.close()
    except Exception:
        pass
    gc.collect()
    time.sleep(0.5)


def _remove_temp_file(path: str) -> None:
    """Remove temp file with retry."""
    if not path or not os.path.exists(path):
        return
    for _ in range(20):
        try:
            os.remove(path)
            return
        except OSError:
            time.sleep(0.3)


def _copy_for_parse(dest_path: str) -> bool:
    """Copy BONEPILE_UPLOAD_PATH to dest_path. Returns True on success."""
    try:
        shutil.copy2(BONEPILE_UPLOAD_PATH, dest_path)
        return True
    except Exception:
        return False


def _save_uploaded_bonepile_file(file_storage) -> Dict[str, Any]:
    ensure_dirs()
    tmp_path = BONEPILE_UPLOAD_PATH + ".tmp"
    file_storage.save(tmp_path)
    dest = BONEPILE_UPLOAD_PATH
    for attempt in range(5):
        try:
            if os.path.exists(dest):
                os.remove(dest)
            os.replace(tmp_path, dest)
            break
        except OSError as e:
            if attempt == 4:
                raise RuntimeError(
                    "Could not replace file. Close any app that has bonepile_upload.xlsx open, then try again. " + str(e)
                ) from e
            time.sleep(0.3 * (attempt + 1))
    stat = os.stat(BONEPILE_UPLOAD_PATH)
    now = datetime.now(CA_TZ).replace(microsecond=0)
    return {
        "has_file": True,
        "path": BONEPILE_UPLOAD_PATH,
        "original_name": getattr(file_storage, "filename", None),
        "size_bytes": int(getattr(stat, "st_size", 0)),
        "uploaded_at_ca_ms": utc_ms(now),
    }


def run_bonepile_parse_job(job_id: str, sheets: Optional[List[str]] = None, path: Optional[str] = None) -> None:
    """Parse the uploaded NV/IGS workbook."""
    parse_path = path
    if not parse_path:
        if not os.path.exists(BONEPILE_UPLOAD_PATH):
            raise RuntimeError("No uploaded bonepile workbook found")
        parse_path = os.path.join(ANALYTICS_CACHE_DIR, "bonepile_parse_" + job_id + ".xlsx")
        if not _copy_for_parse(parse_path):
            raise RuntimeError("Could not copy workbook for parse (file may be in use)")
    try:
        set_job(job_id, status="running", message="Parsing workbook...", started_at=int(time.time()))
        ensure_db_ready()
        with scan_lock:
            state = RawState.load()
        if not os.path.exists(parse_path):
            raise RuntimeError("No uploaded bonepile workbook found")
        wb = None
        wb = _load_bonepile_workbook(parse_path)
        try:
            all_sheets = list(wb.sheetnames)
            allowed = [s for s in BONEPILE_ALLOWED_SHEETS if s in all_sheets]
            target = allowed if not sheets else [s for s in sheets if s in allowed]

            mapping_cfg = (state.bonepile_mapping or {})
            sheet_status: Dict[str, Any] = state.bonepile_sheet_status or {}
            bp_sns_this_run: Set[str] = set()

            conn = connect_db()
            try:
                for sheet in target:
                    ws = wb[sheet]
                    
                    current_hash = _hash_sheet_content(ws)
                    prev_status = sheet_status.get(sheet) if isinstance(sheet_status.get(sheet), dict) else {}
                    prev_hash = prev_status.get("content_hash") if isinstance(prev_status.get("content_hash"), str) else None
                    
                    if prev_hash and prev_hash == current_hash:
                        prev_status["last_run_ca_ms"] = utc_ms(datetime.now(CA_TZ))
                        prev_status["skipped"] = True
                        prev_status["skip_reason"] = "Content unchanged (hash match)"
                        sheet_status[sheet] = prev_status
                        continue
                    
                    cfg = (mapping_cfg.get(sheet) or {}) if isinstance(mapping_cfg.get(sheet), dict) else {}
                    header_row = int(cfg.get("header_row") or 0) if cfg.get("header_row") else 0
                    if header_row <= 0:
                        header_row = _find_header_row(ws) or 0
                    if header_row <= 0:
                        sheet_status[sheet] = {
                            "status": "error",
                            "error": "Header row not found (SN)",
                            "last_run_ca_ms": utc_ms(datetime.now(CA_TZ)),
                            "content_hash": current_hash,
                        }
                        continue

                    header_map = _read_header_map(ws, header_row=header_row)

                    col_map: Dict[str, int] = {}
                    user_cols = cfg.get("columns") if isinstance(cfg.get("columns"), dict) else None
                    if user_cols:
                        for k, v in user_cols.items():
                            if not v:
                                continue
                            if isinstance(v, str):
                                col_map[k] = int(header_map.get(v.strip().upper(), 0))
                            else:
                                try:
                                    col_map[k] = int(v)
                                except Exception:
                                    col_map[k] = 0
                        auto = _auto_mapping_from_headers(header_map)
                        for k, idx in auto.items():
                            col_map.setdefault(k, idx)
                    else:
                        col_map = _auto_mapping_from_headers(header_map)

                    errs = _mapping_errors(col_map, header_map)
                    if errs:
                        sheet_status[sheet] = {
                            "status": "error",
                            "error": "; ".join(errs[:3]),
                            "header_row": header_row,
                            "last_run_ca_ms": utc_ms(datetime.now(CA_TZ)),
                            "content_hash": current_hash,
                        }
                        continue

                    conn.execute("DELETE FROM bonepile_entries WHERE sheet = ?;", (sheet,))
                    now_ms = utc_ms(datetime.now(CA_TZ).replace(microsecond=0))
                    inserted = 0
                    empty_sn_streak = 0

                    for excel_row_idx, row in enumerate(
                        ws.iter_rows(min_row=header_row + 1, values_only=True),
                        start=header_row + 1,
                    ):
                        if row is None:
                            continue
                        sn_val = row[col_map["sn"] - 1] if col_map["sn"] > 0 and col_map["sn"] <= len(row) else None
                        sn = _normalize_sn(sn_val)
                        if not sn:
                            empty_sn_streak += 1
                            if empty_sn_streak >= 200:
                                break
                            continue
                        empty_sn_streak = 0
                        bp_sns_this_run.add(sn)

                        def cell(idx: int) -> str:
                            if idx <= 0 or idx > len(row):
                                return ""
                            v = row[idx - 1]
                            return "" if v is None else str(v).strip()

                        nv_dispo = cell(col_map.get("nv_disposition", 0))
                        igs_action = cell(col_map.get("igs_action", 0))
                        status = cell(col_map.get("status", 0))
                        pic = cell(col_map.get("pic", 0))
                        igs_status = cell(col_map.get("igs_status", 0))
                        nvpn = cell(col_map.get("nvpn", 0))

                        nv_cnt = len(_extract_mmdd_entries(nv_dispo))
                        igs_cnt = len(_extract_mmdd_entries(igs_action))

                        conn.execute(
                            """
                            INSERT OR REPLACE INTO bonepile_entries (
                              sheet, excel_row, sn, nvpn, status, pic, igs_status,
                              nv_disposition, igs_action, nv_dispo_count, igs_action_count, updated_at_ca_ms
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
                            """,
                            (
                                sheet,
                                int(excel_row_idx),
                                sn,
                                nvpn,
                                status,
                                pic,
                                igs_status,
                                nv_dispo,
                                igs_action,
                                int(nv_cnt),
                                int(igs_cnt),
                                int(now_ms),
                            ),
                        )
                        inserted += 1

                    conn.commit()
                    sheet_status[sheet] = {
                        "status": "ok",
                        "rows": int(inserted),
                        "header_row": int(header_row),
                        "last_run_ca_ms": int(now_ms),
                        "content_hash": current_hash,
                    }

                with scan_lock:
                    st = RawState.load()
                    st.bonepile_sheet_status = sheet_status
                    st.save()
            finally:
                conn.close()
        finally:
            if wb is not None:
                _close_and_release_workbook(wb)
                wb = None
            _remove_temp_file(parse_path)

        if bp_sns_this_run:
            update_bp_sn_cache(bp_sns_this_run)

        set_job(job_id, status="done", message="Workbook parsed", finished_at=int(time.time()))
    except Exception as e:
        set_job(job_id, status="error", error=str(e), finished_at=int(time.time()))
        _remove_temp_file(parse_path)
        with scan_lock:
            st = RawState.load()
            ss = st.bonepile_sheet_status or {}
            ss["_job_error"] = str(e)
            st.bonepile_sheet_status = ss
            st.save()


def _bonepile_status_payload(state: RawState) -> Dict[str, Any]:
    bf = state.bonepile_file or {}
    return {
        "file": bf,
        "allowed_sheets": BONEPILE_ALLOWED_SHEETS,
        "mapping": state.bonepile_mapping or {},
        "sheets": state.bonepile_sheet_status or {},
    }


def _parse_ca_input_datetime(s: str, *, is_end: bool) -> Optional[datetime]:
    """Parse user-provided datetime string in CA timezone."""
    if not s:
        return None
    s = str(s).strip()
    try:
        if re.search(r"\d:\d\d:\d\d$", s):
            dt = datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
            return CA_TZ.localize(dt)
        dt = datetime.strptime(s, "%Y-%m-%d %H:%M")
        dt_ca = CA_TZ.localize(dt)
        if is_end:
            dt_ca = dt_ca + timedelta(seconds=59)
        return dt_ca
    except Exception:
        return None


def _last_mmdd_only(text: Any) -> Optional[Tuple[int, int]]:
    """Return (month, day) from the last mm/dd in cell text, or None."""
    raw = (str(text) if text is not None else "").strip()
    if not raw:
        return None
    matches = list(re.finditer(r"\b(\d{1,2})/(\d{1,2})\b", raw))
    if not matches:
        return None
    m = matches[-1]
    try:
        month = int(m.group(1))
        day = int(m.group(2))
        if 1 <= month <= 12 and 1 <= day <= 31:
            return (month, day)
    except (ValueError, IndexError):
        pass
    return None


def _disposition_period_from_row(
    row: Dict[str, Any], aggregation: str, fallback_ca_ms: Optional[int] = None
) -> str:
    """Return period key from row: use last mm/dd in nv_disposition or igs_action."""
    ca_ms = row.get("updated_at_ca_ms") or fallback_ca_ms
    year = None
    if ca_ms is not None:
        try:
            dt = datetime.fromtimestamp(ca_ms / 1000.0, tz=CA_TZ)
            year = dt.year
        except Exception:
            pass
    mmdd_nv = _last_mmdd_only(row.get("nv_disposition"))
    mmdd_igs = _last_mmdd_only(row.get("igs_action"))
    mmdd = mmdd_igs or mmdd_nv
    if mmdd is not None and year is not None:
        try:
            d = date(year, mmdd[0], mmdd[1])
            if aggregation == "monthly":
                return d.strftime("%Y-%m")
            if aggregation == "weekly":
                days_since_sunday = (d.weekday() + 1) % 7
                week_start = d - timedelta(days=days_since_sunday)
                week_end = week_start + timedelta(days=6)
                return f"{week_start.strftime('%Y-%m-%d')}~{week_end.strftime('%Y-%m-%d')}"
            return d.strftime("%Y-%m-%d")
        except (ValueError, TypeError):
            pass
    return _disposition_period_from_ca_ms(ca_ms, aggregation)


def _disposition_period_from_ca_ms(ca_ms: Optional[int], aggregation: str) -> str:
    """Return period key from updated_at_ca_ms."""
    if ca_ms is None:
        return ""
    try:
        dt = datetime.fromtimestamp(ca_ms / 1000.0, tz=CA_TZ)
    except Exception:
        return ""
    if aggregation == "weekly":
        days_since_sunday = (dt.weekday() + 1) % 7
        week_start = (dt - timedelta(days=days_since_sunday)).date()
        week_end = week_start + timedelta(days=6)
        return f"{week_start.strftime('%Y-%m-%d')}~{week_end.strftime('%Y-%m-%d')}"
    if aggregation == "monthly":
        return dt.strftime("%Y-%m")
    return dt.strftime("%Y-%m-%d")


def _is_pass_status(status_norm: str) -> bool:
    """Check if normalized status string indicates a pass status."""
    if not status_norm:
        return False
    return (
        "PASS" in status_norm or
        "ALL PASS" in status_norm or
        "PASS ALL" in status_norm or
        "PASSED" in status_norm
    )


def _last_mmdd_entry(text: Any) -> str:
    """Return the last mm/dd entry segment from cell text, or full text if no mm/dd."""
    entries = _extract_mmdd_entries(text)
    if entries:
        return entries[-1].strip()
    return (str(text) if text is not None else "").strip()


def _last_entry_for_mmdd(text: Any, month: int, day: int) -> str:
    """Return the last disposition entry segment that has the given mm/dd, or empty string."""
    entries = _extract_mmdd_entries(text)
    for i in range(len(entries) - 1, -1, -1):
        seg = entries[i]
        m = re.search(r"\b(\d{1,2})/(\d{1,2})\b", seg)
        if m:
            try:
                if int(m.group(1)) == month and int(m.group(2)) == day:
                    return seg.strip()
            except (ValueError, TypeError):
                pass
    return ""


def compute_disposition_stats(aggregation: str = "daily", start_ca_ms: Optional[int] = None, end_ca_ms: Optional[int] = None) -> Dict[str, Any]:
    """
    Compute NV Disposition stats from bonepile_entries.
    Only considers SNs where Status=FAIL and PIC=IGS.
    Logic:
    - Waiting: igs_action empty OR last date in IGS action < last date in NV disposition
    - Complete: last date in IGS action >= last date in NV disposition
    - Total: waiting + complete
    """
    conn = connect_db()
    try:
        query = "SELECT sn, nvpn, status, pic, nv_disposition, igs_action, updated_at_ca_ms FROM bonepile_entries;"
        rows = conn.execute(query).fetchall()
    finally:
        conn.close()

    def _norm(s: Any) -> str:
        return (str(s) if s is not None else "").strip().upper()

    def _row_dict(r) -> Dict[str, Any]:
        return {k: r[k] for k in r.keys()} if hasattr(r, "keys") else dict(r)

    if start_ca_ms is not None and end_ca_ms is not None:
        start_d = datetime.fromtimestamp(start_ca_ms / 1000.0, tz=CA_TZ).date()
        end_d = datetime.fromtimestamp(end_ca_ms / 1000.0, tz=CA_TZ).date()
        year = start_d.year
    else:
        start_d = None
        end_d = None
        year = datetime.now(CA_TZ).year

    def _date_to_period(d: date) -> str:
        if aggregation == "monthly":
            return d.strftime("%Y-%m")
        elif aggregation == "weekly":
            days_since_sunday = (d.weekday() + 1) % 7
            week_start = d - timedelta(days=days_since_sunday)
            week_end = week_start + timedelta(days=6)
            return f"{week_start.strftime('%Y-%m-%d')}~{week_end.strftime('%Y-%m-%d')}"
        else:
            return d.strftime("%Y-%m-%d")

    # Per SN, keep one row; tie-break by latest last mm/dd in nv_disposition
    sn_latest: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        rd = _row_dict(r)
        sn = (rd.get("sn") or "").strip()
        if not sn:
            continue
        ca_ms = rd.get("updated_at_ca_ms")
        nv_text = rd.get("nv_disposition")
        last_mmdd = _last_mmdd_only(nv_text) if nv_text else None
        existing = sn_latest.get(sn)
        if existing is None:
            sn_latest[sn] = {"row": r, "rd": rd, "updated_at_ca_ms": ca_ms, "last_mmdd": last_mmdd}
            continue
        existing_ca = existing.get("updated_at_ca_ms") or 0
        if (ca_ms or 0) > existing_ca:
            sn_latest[sn] = {"row": r, "rd": rd, "updated_at_ca_ms": ca_ms, "last_mmdd": last_mmdd}
            continue
        if (ca_ms or 0) == existing_ca and last_mmdd and existing.get("last_mmdd"):
            try:
                cur_d = date(year, last_mmdd[0], last_mmdd[1])
                exist_d = date(year, existing["last_mmdd"][0], existing["last_mmdd"][1])
                if cur_d > exist_d:
                    sn_latest[sn] = {"row": r, "rd": rd, "updated_at_ca_ms": ca_ms, "last_mmdd": last_mmdd}
            except (ValueError, TypeError):
                pass

    # New logic: Only Status=FAIL & PIC=IGS. Waiting = igs empty or igs_date < nv_date; Complete = igs_date >= nv_date
    waiting_sns: Dict[str, Dict[str, Any]] = {}
    complete_sns: Dict[str, Dict[str, Any]] = {}
    for sn, data in sn_latest.items():
        rd = data["rd"]
        if _norm(rd.get("status")) != "FAIL" or _norm(rd.get("pic")) != "IGS":
            continue
        mmdd_nv = _last_mmdd_only(rd.get("nv_disposition"))
        if mmdd_nv is None:
            continue
        try:
            nv_date = date(year, mmdd_nv[0], mmdd_nv[1])
            if start_d and nv_date < start_d - timedelta(days=60):
                nv_date = date(year + 1, mmdd_nv[0], mmdd_nv[1])
        except (ValueError, TypeError):
            continue
        if start_d is not None and end_d is not None:
            if not (start_d <= nv_date <= end_d):
                continue
        sku = (rd.get("nvpn") or "").strip() or "Unknown"
        period_nv = _date_to_period(nv_date)
        mmdd_igs = _last_mmdd_only(rd.get("igs_action"))
        if mmdd_igs is None:
            waiting_sns[sn] = {"sku": sku, "period_nv": period_nv}
        else:
            try:
                igs_date = date(year, mmdd_igs[0], mmdd_igs[1])
                if start_d and igs_date < start_d - timedelta(days=60):
                    igs_date = date(year + 1, mmdd_igs[0], mmdd_igs[1])
            except (ValueError, TypeError):
                waiting_sns[sn] = {"sku": sku, "period_nv": period_nv}
                continue
            if igs_date >= nv_date:
                complete_sns[sn] = {"sku": sku, "period_nv": period_nv}
            else:
                waiting_sns[sn] = {"sku": sku, "period_nv": period_nv}

    # Build by_period and by_sku
    by_period: Dict[str, Dict[str, Any]] = {}
    by_sku: Dict[str, Dict[str, Any]] = {}

    for sn, info in waiting_sns.items():
        p = info["period_nv"]
        by_period.setdefault(p, {"period": p, "total": 0, "waiting_igs": 0, "complete": 0})
        by_period[p]["waiting_igs"] += 1
        by_period[p]["total"] += 1
        sku = info["sku"]
        by_sku.setdefault(sku, {"sku": sku, "total": 0, "waiting_igs": 0, "complete": 0})
        by_sku[sku]["waiting_igs"] += 1
        by_sku[sku]["total"] += 1
    for sn, info in complete_sns.items():
        p = info["period_nv"]
        by_period.setdefault(p, {"period": p, "total": 0, "waiting_igs": 0, "complete": 0})
        by_period[p]["complete"] += 1
        by_period[p]["total"] += 1
        sku = info["sku"]
        by_sku.setdefault(sku, {"sku": sku, "total": 0, "waiting_igs": 0, "complete": 0})
        by_sku[sku]["complete"] += 1
        by_sku[sku]["total"] += 1

    summary_waiting = len(waiting_sns)
    summary_complete = len(complete_sns)
    summary_total = summary_waiting + summary_complete
    summary = {
        "total": summary_total,
        "waiting_igs": summary_waiting,
        "complete": summary_complete,
    }

    # Count unique trays (SNs) in BP and trays with ALL PASS status
    conn_all = connect_db()
    try:
        all_rows = conn_all.execute(
            "SELECT sn, nvpn, status, updated_at_ca_ms FROM bonepile_entries;"
        ).fetchall()
    finally:
        conn_all.close()
    
    sn_latest_row: Dict[str, Dict[str, Any]] = {}
    for r in all_rows:
        sn = (r["sn"] or "").strip()
        if not sn:
            continue
        ca_ms = r["updated_at_ca_ms"] or 0
        existing = sn_latest_row.get(sn)
        if existing is None or ca_ms > (existing.get("updated_at_ca_ms") or 0):
            sn_latest_row[sn] = {
                "sn": sn,
                "nvpn": (r["nvpn"] or "").strip() or "Unknown",
                "status": (r["status"] or "").strip(),
                "updated_at_ca_ms": ca_ms,
            }
    
    unique_trays_bp = len(sn_latest_row)
    all_pass_trays = 0
    tray_by_sku: Dict[str, Dict[str, int]] = {}
    
    for sn, d in sn_latest_row.items():
        status_norm = _norm(d["status"])
        if _is_pass_status(status_norm):
            all_pass_trays += 1
        
        sku = d["nvpn"]
        tray_by_sku.setdefault(sku, {"sku": sku, "total_trays": 0, "all_pass_trays": 0})
        tray_by_sku[sku]["total_trays"] += 1
        if _is_pass_status(status_norm):
            tray_by_sku[sku]["all_pass_trays"] += 1
    
    summary["unique_trays_bp"] = unique_trays_bp
    summary["all_pass_trays"] = all_pass_trays
    tray_by_sku_list = sorted(tray_by_sku.values(), key=lambda x: (x["sku"]))

    by_sku_list = sorted(by_sku.values(), key=lambda x: (x["sku"]))
    by_period_list = sorted(by_period.values(), key=lambda x: (x["period"]))

    # Filter by_period to only include periods within user's date range
    if start_ca_ms is not None and end_ca_ms is not None:
        try:
            start_d = datetime.fromtimestamp(start_ca_ms / 1000.0, tz=CA_TZ).date()
            end_d = datetime.fromtimestamp(end_ca_ms / 1000.0, tz=CA_TZ).date()
            filtered = []
            for p in by_period_list:
                period_str = p.get("period") or ""
                if aggregation == "daily" and re.match(r"^\d{4}-\d{2}-\d{2}$", period_str):
                    pd = datetime.strptime(period_str, "%Y-%m-%d").date()
                    if start_d <= pd <= end_d:
                        filtered.append(p)
                elif aggregation == "monthly" and re.match(r"^\d{4}-\d{2}$", period_str):
                    pd = datetime.strptime(period_str + "-01", "%Y-%m-%d").date()
                    if start_d <= pd <= end_d:
                        filtered.append(p)
                elif aggregation == "weekly" and "~" in period_str:
                    part = period_str.split("~")[0]
                    if re.match(r"^\d{4}-\d{2}-\d{2}$", part):
                        pd = datetime.strptime(part, "%Y-%m-%d").date()
                        if pd <= end_d and (pd + timedelta(days=6)) >= start_d:
                            filtered.append(p)
                else:
                    filtered.append(p)
            by_period_list = filtered
        except Exception:
            pass

    return {"summary": summary, "by_sku": by_sku_list, "by_period": by_period_list, "tray_by_sku": tray_by_sku_list}


def compute_disposition_sn_list(
    metric: str, sku: Optional[str] = None, period: Optional[str] = None, aggregation: str = "daily", start_ca_ms: Optional[int] = None, end_ca_ms: Optional[int] = None
) -> List[Dict[str, Any]]:
    """Return list of { sn, last_nv_dispo, last_igs_action, nvpn, status, pic } for drill-down."""
    def _norm(s: Any) -> str:
        return (str(s) if s is not None else "").strip().upper()

    def _row_dict(r) -> Dict[str, Any]:
        return {k: r[k] for k in r.keys()} if hasattr(r, "keys") else dict(r)

    # Special handling for trays_bp and all_pass_trays
    if metric in ("trays_bp", "all_pass_trays"):
        conn = connect_db()
        try:
            rows = conn.execute(
                "SELECT sheet, excel_row, sn, nvpn, status, pic, nv_disposition, igs_action, updated_at_ca_ms FROM bonepile_entries;"
            ).fetchall()
        finally:
            conn.close()
        
        sn_rows: Dict[str, Dict[str, Any]] = {}
        for r in rows:
            rd = _row_dict(r)
            sn = (rd.get("sn") or "").strip()
            if not sn:
                continue
            ca_ms = rd.get("updated_at_ca_ms")
            existing = sn_rows.get(sn)
            if existing is None or (ca_ms or 0) > (existing.get("updated_at_ca_ms") or 0):
                sn_rows[sn] = {
                    "sn": sn,
                    "nvpn": (rd.get("nvpn") or "").strip() or "Unknown",
                    "status": (rd.get("status") or "").strip(),
                    "pic": (rd.get("pic") or "").strip(),
                    "nv_disposition": rd.get("nv_disposition"),
                    "igs_action": rd.get("igs_action"),
                    "updated_at_ca_ms": ca_ms,
                }
        
        out: List[Dict[str, Any]] = []
        for sn, d in sn_rows.items():
            status_norm = _norm(d.get("status"))
            row_sku = d.get("nvpn")
            if sku and sku != "__TOTAL__" and row_sku != sku:
                continue
            if metric == "all_pass_trays" and not _is_pass_status(status_norm):
                continue
            out.append({
                "sn": sn,
                "last_nv_dispo": _last_mmdd_entry(d.get("nv_disposition")),
                "last_igs_action": _last_mmdd_entry(d.get("igs_action")),
                "nvpn": row_sku,
                "status": d.get("status"),
                "pic": d.get("pic"),
            })
        out.sort(key=lambda x: (x["sn"]))
        return out

    # waiting | complete | total - new logic: Status=FAIL & PIC=IGS; waiting = igs empty or igs_date < nv_date; complete = igs_date >= nv_date
    def _date_to_period(d: date) -> str:
        if aggregation == "monthly":
            return d.strftime("%Y-%m")
        elif aggregation == "weekly":
            days_since_sunday = (d.weekday() + 1) % 7
            week_start = d - timedelta(days=days_since_sunday)
            week_end = week_start + timedelta(days=6)
            return f"{week_start.strftime('%Y-%m-%d')}~{week_end.strftime('%Y-%m-%d')}"
        return d.strftime("%Y-%m-%d")

    conn = connect_db()
    try:
        rows = conn.execute(
            "SELECT sheet, excel_row, sn, nvpn, status, pic, nv_disposition, igs_action, updated_at_ca_ms FROM bonepile_entries;"
        ).fetchall()
    finally:
        conn.close()

    year = datetime.now(CA_TZ).year
    if start_ca_ms is not None and end_ca_ms is not None:
        start_d = datetime.fromtimestamp(start_ca_ms / 1000.0, tz=CA_TZ).date()
        end_d = datetime.fromtimestamp(end_ca_ms / 1000.0, tz=CA_TZ).date()
        year = start_d.year
    else:
        start_d = None
        end_d = None

    sn_data: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        rd = _row_dict(r)
        sn = (rd.get("sn") or "").strip()
        if not sn:
            continue
        ca_ms = rd.get("updated_at_ca_ms")
        last_mmdd = _last_mmdd_only(rd.get("nv_disposition"))
        existing = sn_data.get(sn)
        if existing is None:
            sn_data[sn] = {"row": r, "rd": rd, "updated_at_ca_ms": ca_ms, "last_mmdd": last_mmdd}
            continue
        if (ca_ms or 0) > (existing.get("updated_at_ca_ms") or 0):
            sn_data[sn] = {"row": r, "rd": rd, "updated_at_ca_ms": ca_ms, "last_mmdd": last_mmdd}
            continue
        if (ca_ms or 0) == (existing.get("updated_at_ca_ms") or 0) and last_mmdd and existing.get("last_mmdd"):
            try:
                cur_d = date(year, last_mmdd[0], last_mmdd[1])
                exist_d = date(year, existing["last_mmdd"][0], existing["last_mmdd"][1])
                if cur_d > exist_d:
                    sn_data[sn] = {"row": r, "rd": rd, "updated_at_ca_ms": ca_ms, "last_mmdd": last_mmdd}
            except (ValueError, TypeError):
                pass

    out: List[Dict[str, Any]] = []
    for sn, data in sn_data.items():
        rd = data["rd"]
        if _norm(rd.get("status")) != "FAIL" or _norm(rd.get("pic")) != "IGS":
            continue
        mmdd_nv = _last_mmdd_only(rd.get("nv_disposition"))
        if mmdd_nv is None:
            continue
        try:
            nv_date = date(year, mmdd_nv[0], mmdd_nv[1])
            if start_d and nv_date < start_d - timedelta(days=60):
                nv_date = date(year + 1, mmdd_nv[0], mmdd_nv[1])
        except (ValueError, TypeError):
            continue
        if start_d is not None and end_d is not None:
            if not (start_d <= nv_date <= end_d):
                continue
        row_sku = (rd.get("nvpn") or "").strip() or "Unknown"
        if sku and sku != "__TOTAL__" and row_sku != sku:
            continue
        period_nv = _date_to_period(nv_date)
        if period and period != "__TOTAL__" and period_nv != period:
            continue
        mmdd_igs = _last_mmdd_only(rd.get("igs_action"))
        is_waiting = False
        if mmdd_igs is None:
            is_waiting = True
        else:
            try:
                igs_date = date(year, mmdd_igs[0], mmdd_igs[1])
                if start_d and igs_date < start_d - timedelta(days=60):
                    igs_date = date(year + 1, mmdd_igs[0], mmdd_igs[1])
                is_waiting = igs_date < nv_date
            except (ValueError, TypeError):
                is_waiting = True
        if metric == "waiting" and not is_waiting:
            continue
        if metric == "complete" and is_waiting:
            continue
        out.append({
            "sn": sn,
            "last_nv_dispo": _last_mmdd_entry(rd.get("nv_disposition")),
            "last_igs_action": _last_mmdd_entry(rd.get("igs_action")),
            "nvpn": row_sku,
            "status": (rd.get("status") or "").strip(),
            "pic": (rd.get("pic") or "").strip(),
        })
    out.sort(key=lambda x: (x["sn"]))
    return out
