# -*- coding: utf-8 -*-
"""Update Tray_Summary_Template.xlsx: SN list headers, formatting (centered, borders, colors).
When export adds rows below, copy style from row 3."""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
except ImportError:
    print("openpyxl required: pip install openpyxl")
    sys.exit(1)

APP_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_PATH = os.path.join(APP_DIR, "templates", "Tray_Summary_Template.xlsx")

SN_HEADERS = ["SN", "BP", "RESULT", "PART_NUMBER", "LAST_STATION", "LAST_TEST_TIME", "ERROR_CODE", "FAILURE_MSG"]

THIN = Side(style="thin", color="000000")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def apply_style(cell, fill=None, alignment=CENTER, border=BORDER_ALL):
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if fill:
        cell.fill = fill


def main():
    if not os.path.isfile(TEMPLATE_PATH):
        print(f"Template not found: {TEMPLATE_PATH}")
        sys.exit(1)
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # Summary table A2:D5 - centered, borders
    for r in range(2, 6):
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            if r == 2 or c == 1:
                apply_style(cell, fill=YELLOW_FILL)
            else:
                apply_style(cell, fill=LIGHT_GREEN_FILL)

    # SN list headers F2:M2 - centered, borders, light blue
    for c, h in enumerate(SN_HEADERS, start=6):
        cell = ws.cell(row=2, column=c, value=h)
        apply_style(cell, fill=LIGHT_BLUE_FILL)
        cell.font = Font(bold=True)

    # SN list data row 3 (template row for copy) - centered, borders, yellow
    for c in range(6, 6 + len(SN_HEADERS)):
        cell = ws.cell(row=3, column=c)
        apply_style(cell, fill=YELLOW_FILL)

    wb.save(TEMPLATE_PATH)
    print(f"Updated {TEMPLATE_PATH}: SN headers, centered alignment, borders, colors. Row 3 is format template for new rows.")


if __name__ == "__main__":
    main()
